from flask import Flask, render_template, request
from openpyxl import Workbook, load_workbook
import cv2
import os

app = Flask(__name__)

student_names = ["BRB_SMBS1", "Pranav_SMBS2", "John_Doe", "Jane_Doe"]
fee_statuses = [1, 0, 1, 1]
boarding_times = ["07:30:00", "07:45:00", "07:50:00", "07:55:00"]

# Load existing student data from the Excel file
try:
    workbook = load_workbook('student_data.xlsx')
    sheet = workbook.active
except FileNotFoundError:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Name", "Fee Status", "Boarding Time"])

@app.route('/')
def index():
    return render_template('index.html', student_names=student_names, fee_statuses=fee_statuses, boarding_times=boarding_times)

@app.route('/update', methods=['POST'])
def update():
    global workbook, sheet

    if request.method == 'POST':
        name = request.form['name']
        fee_status = int(request.form['fee_status'])
        boarding_time = request.form['boarding_time']

        # Append the entered data to the Excel sheet and arrays
        sheet.append([name, fee_status, boarding_time])
        student_names.append(name)
        fee_statuses.append(fee_status)
        boarding_times.append(boarding_time)

        # Save the changes to the Excel file
        workbook.save('student_data.xlsx')

    return render_template('index.html', student_names=student_names, fee_statuses=fee_statuses, boarding_times=boarding_times)

@app.route('/capture_photos', methods=['POST'])
def capture_photos():
    if request.method == 'POST':
        name_photos = request.form['name_photos']
        directory = f'C:/Users/bvspranav999/OneDrive/Desktop/Face Recognition Project/Real-time-face-recognition-Using-Facenet-main/Faces/{name_photos}'

        # Create a directory if it doesn't exist
        if not os.path.exists(directory):
            os.makedirs(directory)

        # Capture 20 photos
        cap = cv2.VideoCapture(0)
        for i in range(20):
            ret, frame = cap.read()
            if not ret:
                break
            cv2.imwrite(f'{directory}/image_{i + 1}.jpg', frame)
        cap.release()

    return render_template('index.html', student_names=student_names, fee_statuses=fee_statuses, boarding_times=boarding_times)

if __name__ == '__main__':
    app.run(debug=True)
