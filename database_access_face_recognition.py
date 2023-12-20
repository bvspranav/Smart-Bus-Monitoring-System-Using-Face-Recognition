
import cv2
import numpy as np
import mtcnn
from architecture import *
from train_v2 import normalize, l2_normalizer
from scipy.spatial.distance import cosine
from tensorflow.keras.models import load_model
import pickle
import telepot
import time
from datetime import datetime, timedelta
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import Workbook, load_workbook

now = datetime.now()
current_time = now.strftime("%H:%M:%S")

# Load the existing Excel file or create a new one
try:
    workbook = load_workbook('student_data.xlsx')
except FileNotFoundError:
    workbook = Workbook()

# Access the active sheet
sheet = workbook.active

# Read existing data from the sheet
student_names = []
fee_statuses = []
boarding_times = []

for row in sheet.iter_rows(min_row=2, values_only=True):
    student_names.append(row[0])
    fee_statuses.append(row[1])
    boarding_times.append(row[2])

print("Student Names:", student_names)
print("Fee Statuses:", fee_statuses)
print("Boarding Times:", boarding_times)

workbook.save('student_data.xlsx') ## if automatically excel is not created, create an excel sheet manually in the same directory


# Google Sheets integration
scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
creds = ServiceAccountCredentials.from_json_keyfile_name('json file path', scope) ## add json file path
client = gspread.authorize(creds)
sheet_name = 'VIT-AP BUS MONITORING SYSTEM'
sheet = client.open(sheet_name).sheet1

receiveTelegramMessage = False

def handle(msg):
    global telegramText
    global chat_id
    global receiveTelegramMessage

    chat_id = msg['chat']['id']
    telegramText = msg['text']

    print("Message received from " + str(chat_id))

    if telegramText == "/start":
        bot.sendMessage(chat_id, "Welcome")
    else:
        receiveTelegramMessage = True

def capture_and_send_to_sheet(name, fee_status, boarding_message):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    image_filename = f'captured_image_{timestamp}.jpg'
    cv2.imwrite(image_filename, frame)

    # Compose a message with details
    message = {
        "Person Name": name,
        "Current Time": current_time,
        "Fee Status": fee_status,
        "Boarding Check": boarding_message
    }

    # Insert data into the Google Sheet

    try:
    # Your Google Sheets operations here
        sheet.append_row(list(message.values()))
        print("Details sent to Google Sheet")
    except Exception as e:
        print(f"Error sending details to Google Sheet: {e}")

def capture_and_send_photo(name, fee_status, boarding_message):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    image_filename = f'captured_image_{timestamp}.jpg'
    cv2.imwrite(image_filename, frame)

    # Compose a message with details
    message = f"Person Name: {name}\n" \
              f"Current Time: {current_time}\n" \
              f"Fee Status: {fee_status}\n" \
              f"Boarding Check: {boarding_message}"

    print("Sending photo and details to " + str(chat_id))

    # Send the photo and details to the Telegram chat
    bot.sendPhoto(chat_id, photo=open(image_filename, 'rb'))
    bot.sendMessage(chat_id, message)
    bot.sendMessage(chat_id, "Student has boarded bus no: 4")

def fee_checker(name, fee_status):
    if fee_status == 0:
        return f'Comments: "{name}" has not paid the fee.'
    elif fee_status == 1:
        return f'Comments: {name} has paid the fee.'
    else:
        return f'Comments: The person is Unknown.'

def boarding_checker(name, boarding_times):
    student_boarding_time = get_student_boarding_time(name)
    if student_boarding_time is not None:
        # Convert string times to datetime objects
        current_time_obj = datetime.strptime(current_time, "%H:%M:%S")
        
        # Convert student_boarding_time to string
        student_boarding_time_str = str(student_boarding_time)
        
        # Convert student_boarding_time to a datetime object
        student_boarding_time_obj = datetime.strptime(student_boarding_time_str, "%H:%M:%S").time()

        # Get today's date to combine with student_boarding_time_obj
        today_date = datetime.now().date()
        student_boarding_time_combined = datetime.combine(today_date, student_boarding_time_obj)

        time_difference = abs(current_time_obj - student_boarding_time_combined)
        if time_difference <= timedelta(minutes=10):
            return f'Authorized entry for {name}.'
        else:
            if(name == "unknown"):
                return f'This student is Unknown'
            else:
                return f'This student {name} has boarded at a different boarding point.'
    else:
        return f'Boarding time not available for {name}.'


confidence_t = 0.99
recognition_t = 0.5
required_size = (160, 160)



def get_face(img, box):
    x1, y1, width, height = box
    x1, y1 = abs(x1), abs(y1)
    x2, y2 = x1 + width, y1 + height
    face = img[y1:y2, x1:x2]
    return face, (x1, y1), (x2, y2)

def get_encode(face_encoder, face, size):
    face = normalize(face)
    face = cv2.resize(face, size)
    encode = face_encoder.predict(np.expand_dims(face, axis=0))[0]
    return encode

def load_pickle(path):
    with open(path, 'rb') as f:
        encoding_dict = pickle.load(f)
    return encoding_dict



def detect(img, detector, encoder, encoding_dict, student_names, fee_statuses, boarding_times):
    img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
    results = detector.detect_faces(img_rgb)
    
    for res in results:
        if res['confidence'] < confidence_t:
            continue

        face, pt_1, pt_2 = get_face(img_rgb, res['box'])
        encode = get_encode(encoder, face, required_size)
        encode = l2_normalizer.transform(encode.reshape(1, -1))[0]
        name = 'unknown'
        fee_status = -1  # Unknown fee status

        distance = float("inf")
        for i, (db_name, db_encode) in enumerate(encoding_dict.items()):
            dist = cosine(db_encode, encode)
            if dist < recognition_t and dist < distance:
                name = db_name
                distance = dist
                fee_status = fee_statuses[student_names.index(name)]  # Use index from student_names list

        fee_message = fee_checker(name, fee_status)

        # Skip checking boarding_time if fee_status is 0
        if fee_status == 0:
            message = fee_message
            boarding_message = "Student has not paid the fee"
        else:
            boarding_message = boarding_checker(name, boarding_times[student_names.index(name)])  # Use index from student_names list
            message = f'{fee_message}\n {boarding_message}'

        capture_and_send_photo(name, fee_status, boarding_message)
        capture_and_send_to_sheet(name, fee_status, boarding_message)

        cv2.rectangle(img, pt_1, pt_2, (0, 255, 0), 2)
        cv2.putText(img, name + f'__{distance:.2f}', (pt_1[0], pt_1[1] - 5), cv2.FONT_HERSHEY_SIMPLEX, 1,
                    (0, 200, 200), 2)
            
    return img



#boarding_times = ["07:30:00", "07:45:00", "07:50:00", "07:55:00"]

def get_student_boarding_time(name):
    # Implement the logic to get the boarding time for a student
    # You may use a dictionary or a database to store the boarding times for each student
    # Return None if the boarding time is not available
    return boarding_times[student_names.index(name)] if name in student_names else None

if __name__ == "__main__":
    

    required_shape = (160, 160)
    face_encoder = InceptionResNetV2()
    path_m = "add keras.h5 path" ## add keras.h5 file path
    face_encoder.load_weights(path_m)
    encodings_path = 'encodings path' ## add your encodings path
    face_detector = mtcnn.MTCNN()
    encoding_dict = load_pickle(encodings_path)

    bot = telepot.Bot('bot') ## add your bot token number 

    chat_id = 'id' ## add your specified bot id
    bot.message_loop(handle)

    print("Telegram bot is ready")

    bot.sendMessage(chat_id, 'BOT STARTED')
    time.sleep(2)

    #fee_statuses = [1, 0, 1, 1]  # Example fee statuses corresponding to student_names  0 - Student fee unpaid;  1 - Student fee paid 
    #student_names = ["BRB_SMBS1", "Pranav_SMBS2", "John_Doe", "Jane_Doe"]
    
    cap = cv2.VideoCapture(0)

    while cap.isOpened():
        ret, frame = cap.read()

        if not ret:
            print("CAM NOT OPENED")
            break

        frame = detect(frame, face_detector, face_encoder, encoding_dict, student_names, fee_statuses, boarding_times)
        cv2.imshow('camera', frame)

        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    # Release the camera and close OpenCV windows when the script ends
    cap.release()
    cv2.destroyAllWindows()
  


